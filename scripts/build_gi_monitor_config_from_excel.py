#!/usr/bin/env python3
"""Build monitor config from the updated Excel tracker.

This version intentionally relies on column positions in the first worksheet
so it keeps working even when the workbook contains Chinese headers that may be
displayed inconsistently in different terminals.
"""

from __future__ import annotations

import json
import re
import base64
import gzip
from collections import OrderedDict
from pathlib import Path


ROOT = Path(__file__).resolve().parent.parent
SOURCE_XLSX = ROOT / "outputs" / "gi_competitors_extraction_updated.xlsx"
BASE_CONFIG = ROOT / "config" / "gi_monitor_config.json"
OUTPUT_CONFIG = ROOT / "config" / "gi_monitor_config.generated.json"
BUNDLED_CONFIG = ROOT / "config" / "gi_monitor_config.generated.json.gz.b64"

# Expected worksheet layout in the first sheet:
# 0 company, 1 disease, 2 target, 3 product, 4 trial id, 5 CN url, 6 global url
COMPANY_COL = 0
DISEASE_COL = 1
TARGET_COL = 2
PRODUCT_COL = 3
TRIAL_COL = 4
CN_URL_COL = 5
GLOBAL_URL_COL = 6


def normalize_text(value: object) -> str:
    if value is None:
        return ""
    return " ".join(str(value).replace("\xa0", " ").split()).strip()


def split_multi(value: str) -> list[str]:
    raw = normalize_text(value)
    if not raw or raw.upper() == "NA":
        return []
    parts: list[str] = []
    for piece in re.split(r"[|;；\n]+", raw):
        item = normalize_text(piece).rstrip(",")
        if item and item.upper() != "NA" and item not in parts:
            parts.append(item)
    return parts


def load_base_terms() -> dict:
    base = json.loads(BASE_CONFIG.read_text(encoding="utf-8"))
    return {
        "scan_days": int(base.get("scan_days", 3)),
        "event_terms": base.get("event_terms", []),
        "gi_context_terms": base.get("gi_context_terms", []),
        "common_paths": [],
    }


def build_company_rows() -> list[dict]:
    from openpyxl import load_workbook

    wb = load_workbook(SOURCE_XLSX, read_only=True, data_only=True)
    ws = wb.worksheets[0]
    rows = ws.iter_rows(values_only=True)
    next(rows, None)  # skip header

    companies: OrderedDict[str, dict] = OrderedDict()
    for row in rows:
        if not row:
            continue
        company_name = normalize_text(row[COMPANY_COL] if len(row) > COMPANY_COL else "")
        if not company_name:
            continue

        company = companies.setdefault(
            company_name,
            {
                "company": company_name,
                "official_urls": [],
                "diseases": [],
                "targets": [],
                "products": [],
                "trial_ids": [],
                "watch_items": [],
            },
        )

        scalar_fields = (
            ("diseases", DISEASE_COL),
            ("targets", TARGET_COL),
            ("products", PRODUCT_COL),
            ("trial_ids", TRIAL_COL),
        )
        for key, idx in scalar_fields:
            value = normalize_text(row[idx] if len(row) > idx else "")
            if value and value.upper() != "NA" and value not in company[key]:
                company[key].append(value)

        urls = []
        urls.extend(split_multi(normalize_text(row[CN_URL_COL] if len(row) > CN_URL_COL else "")))
        urls.extend(split_multi(normalize_text(row[GLOBAL_URL_COL] if len(row) > GLOBAL_URL_COL else "")))
        for url in urls:
            if url not in company["official_urls"]:
                company["official_urls"].append(url)

        watch_item = {
            "disease": normalize_text(row[DISEASE_COL] if len(row) > DISEASE_COL else ""),
            "target": normalize_text(row[TARGET_COL] if len(row) > TARGET_COL else ""),
            "product": normalize_text(row[PRODUCT_COL] if len(row) > PRODUCT_COL else ""),
            "trial_id": normalize_text(row[TRIAL_COL] if len(row) > TRIAL_COL else ""),
        }
        if watch_item not in company["watch_items"]:
            company["watch_items"].append(watch_item)

    return list(companies.values())


def reuse_existing_generated_config() -> bool:
    if BUNDLED_CONFIG.exists():
        payload = BUNDLED_CONFIG.read_text(encoding="ascii").strip()
        if payload:
            decoded = gzip.decompress(base64.b64decode(payload)).decode("utf-8")
            OUTPUT_CONFIG.write_text(decoded, encoding="utf-8")
            print(f"Excel file not found, restored generated config from bundle: {BUNDLED_CONFIG}")
            return True
    if OUTPUT_CONFIG.exists():
        print(f"Excel file not found, reusing existing generated config: {OUTPUT_CONFIG}")
        return True
    return False


def main() -> None:
    if not SOURCE_XLSX.exists():
        if reuse_existing_generated_config():
            return
        raise FileNotFoundError(
            f"Excel source file not found: {SOURCE_XLSX}. "
            f"Please commit the workbook or provide {OUTPUT_CONFIG} first."
        )

    config = {
        **load_base_terms(),
        "companies": build_company_rows(),
    }
    OUTPUT_CONFIG.write_text(json.dumps(config, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    print(OUTPUT_CONFIG)


if __name__ == "__main__":
    main()
